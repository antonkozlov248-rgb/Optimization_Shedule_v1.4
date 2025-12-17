from flask import Flask, request, jsonify, send_file, abort
from flask_cors import CORS
import json
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from genetic_algorithm import GeneticAlgorithm
from collections import defaultdict

app = Flask(__name__)
CORS(app)

ga_instance = None
last_schedule = None
last_fitness = 0
last_conflicts = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ==================== ЭКСПОРТ В EXCEL ====================

def export_to_excel(schedule, fitness, conflicts):
	"""Экспортировать расписание в XLSX файл с 4 листами"""
	wb = openpyxl.Workbook()
	wb.remove(wb.active)

	header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
	header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
	title_font = Font(name='Calibri', bold=True, size=16, color="4472C4")
	title_14_font = Font(name='Calibri', bold=True, size=14)
	border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
	left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
	alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
	days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']

	# ===== ЛИСТ 0: ОТЧЕТ ГЕНЕРАЦИИ =====
	def create_report_sheet(wb, schedule, fitness, conflicts):
		ws = wb.create_sheet("Отчет генерации", 0)
		ws['A1'] = "ОТЧЕТ ОБ ОПТИМИЗАЦИИ РАСПИСАНИЯ"
		ws['A1'].font = title_font
		ws['A1'].alignment = center_align
		ws.merge_cells('A1:B1')
		ws.row_dimensions[1].height = 25

		total_lessons = sum(len([l for l in lessons if l is not None]) for class_schedule in schedule.values() for lessons in class_schedule.values())

		unique_teachers = set()
		for class_schedule in schedule.values():
			for lessons in class_schedule.values():
				for lesson in lessons:
					if lesson is not None:
						teacher = lesson.get('учитель', '')
						if teacher and teacher != 'Generic' and teacher != 'N/A':
							unique_teachers.add(teacher)

		unique_classes = len(schedule)
		metrics = [
			("Качество расписания (%)", f"{fitness:.1f}"),
			("Всего уроков", str(total_lessons)),
			("Конфликты учителей", str(conflicts.get('teacher_conflicts', 0))),
			("Конфликты кабинетов", str(conflicts.get('room_conflicts', 0))),
			("Нарушения SanPin", str(conflicts.get('sanpin_violations', 0))),
			("Дата создания", datetime.now().strftime("%Y-%m-%d")),
			("Количество классов", str(unique_classes)),
			("Количество учителей", str(len(unique_teachers)))
		]

		ws['A3'] = "Параметр"
		ws['B3'] = "Значение"
		ws['A3'].font = header_font
		ws['B3'].font = header_font
		ws['A3'].fill = header_fill
		ws['B3'].fill = header_fill
		ws['A3'].alignment = center_align
		ws['B3'].alignment = center_align
		ws['A3'].border = border
		ws['B3'].border = border

		row = 4
		for i, (param, value) in enumerate(metrics):
			ws[f'A{row}'] = param
			ws[f'B{row}'] = value
			ws[f'A{row}'].border = border
			ws[f'B{row}'].border = border
			ws[f'A{row}'].alignment = left_align
			ws[f'B{row}'].alignment = center_align
			ws[f'A{row}'].font = Font(name='Calibri', size=11)
			ws[f'B{row}'].font = Font(name='Calibri', size=11)
			if i % 2 == 1:
				ws[f'A{row}'].fill = alt_fill
				ws[f'B{row}'].fill = alt_fill
			row += 1

		ws.column_dimensions['A'].width = 25
		ws.column_dimensions['B'].width = 20
		for r in range(3, row):
			ws.row_dimensions[r].height = 20

	# ===== ЛИСТ 1: РАСПИСАНИЕ КЛАССОВ =====
	def create_classes_sheet(wb, schedule, days):
		ws = wb.create_sheet("Расписание классов", 1)
		row = 1
		sorted_classes = sorted(schedule.keys())

		for class_idx, class_name in enumerate(sorted_classes):
			ws[f'A{row}'] = f"Расписание класса {class_name}"
			ws[f'A{row}'].font = title_14_font
			ws.merge_cells(f'A{row}:H{row}')
			ws.row_dimensions[row].height = 22
			row += 2

			ws[f'A{row}'] = "День"
			ws[f'A{row}'].font = header_font
			ws[f'A{row}'].fill = header_fill
			ws[f'A{row}'].alignment = center_align
			ws[f'A{row}'].border = border
			for period in range(1, 8):
				col = get_column_letter(period + 1)
				ws[f'{col}{row}'] = f"Урок {period}"
				ws[f'{col}{row}'].font = header_font
				ws[f'{col}{row}'].fill = header_fill
				ws[f'{col}{row}'].alignment = center_align
				ws[f'{col}{row}'].border = border
			row += 1

			class_schedule = schedule[class_name]
			for day_idx, day in enumerate(days):
				ws[f'A{row}'] = day
				ws[f'A{row}'].border = border
				ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
				ws[f'A{row}'].alignment = center_align
				if day_idx % 2 == 1:
					ws[f'A{row}'].fill = alt_fill

				lessons = class_schedule.get(day, [])
				if isinstance(lessons, dict):
					lessons = list(lessons.values())

				for period in range(1, 8):
					col = get_column_letter(period + 1)
					target_lesson = None
					for lesson in lessons:
						if lesson is not None:
							lesson_num = lesson.get('урок') if isinstance(lesson, dict) else getattr(lesson, 'урок', 0)
							if lesson_num == period:
								target_lesson = lesson
								break

					if target_lesson:
						subj = target_lesson.get('предмет', '') if isinstance(target_lesson, dict) else getattr(target_lesson, 'предмет', '')
						teacher = target_lesson.get('учитель', '') if isinstance(target_lesson, dict) else getattr(target_lesson, 'учитель', '')
						room = target_lesson.get('кабинет', '') if isinstance(target_lesson, dict) else getattr(target_lesson, 'кабинет', '')
						text = f"{subj}\n{teacher}\n(Каб. {room})"
						ws[f'{col}{row}'] = text
					else:
						ws[f'{col}{row}'] = ""

					ws[f'{col}{row}'].border = border
					ws[f'{col}{row}'].alignment = center_align
					ws[f'{col}{row}'].font = Font(name='Calibri', size=11)
					if day_idx % 2 == 1:
						ws[f'{col}{row}'].fill = alt_fill

				ws.row_dimensions[row].height = 55
				row += 1

			row += 2

		ws.column_dimensions['A'].width = 15
		for col in range(2, 9):
			ws.column_dimensions[get_column_letter(col)].width = 18

	# ===== ЛИСТ 2: РАСПИСАНИЕ УЧИТЕЛЕЙ =====
	def create_teachers_sheet(wb, schedule, days):
		ws = wb.create_sheet("Расписание учителей", 2)
		teachers_set = set()
		teacher_classes = defaultdict(set)

		for class_name, class_schedule in schedule.items():
			for day, lessons in class_schedule.items():
				for lesson in lessons:
					if lesson is not None:
						teacher = lesson.get('учитель', '')
						if teacher and teacher != 'Generic' and teacher != 'N/A':
							teachers_set.add(teacher)
							teacher_classes[teacher].add(class_name)

		sorted_teachers = sorted(teachers_set)
		row = 1

		for teacher in sorted_teachers:
			ws[f'A{row}'] = f"Расписание учителя: {teacher}"
			ws[f'A{row}'].font = title_14_font
			ws.merge_cells(f'A{row}:H{row}')
			ws.row_dimensions[row].height = 22
			row += 2

			ws[f'A{row}'] = "День"
			ws[f'A{row}'].font = header_font
			ws[f'A{row}'].fill = header_fill
			ws[f'A{row}'].alignment = center_align
			ws[f'A{row}'].border = border
			for period in range(1, 8):
				col = get_column_letter(period + 1)
				ws[f'{col}{row}'] = f"Урок {period}"
				ws[f'{col}{row}'].font = header_font
				ws[f'{col}{row}'].fill = header_fill
				ws[f'{col}{row}'].alignment = center_align
				ws[f'{col}{row}'].border = border
			row += 1

			total_hours = 0
			gaps = 0

			for day_idx, day in enumerate(days):
				ws[f'A{row}'] = day
				ws[f'A{row}'].border = border
				ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
				ws[f'A{row}'].alignment = center_align
				if day_idx % 2 == 1:
					ws[f'A{row}'].fill = alt_fill

				day_lessons = []
				for class_name in schedule.keys():
					class_schedule = schedule[class_name]
					lessons = class_schedule.get(day, [])
					if isinstance(lessons, dict):
						lessons = list(lessons.values())
					for lesson in lessons:
						if lesson is not None:
							lesson_teacher = lesson.get('учитель', '')
							if lesson_teacher == teacher:
								day_lessons.append({
									'период': lesson.get('урок', 0),
									'класс': class_name,
									'предмет': lesson.get('предмет', ''),
									'кабинет': lesson.get('кабинет', '')
								})

				if day_lessons:
					day_lessons.sort(key=lambda x: x['период'])
					for i in range(1, 8):
						has_lesson = any(l['период'] == i for l in day_lessons)
						if not has_lesson and i > 0:
							gaps += 1

				for period in range(1, 8):
					col = get_column_letter(period + 1)
					target_lesson = None
					for lesson in day_lessons:
						if lesson['период'] == period:
							target_lesson = lesson
							break

					if target_lesson:
						text = f"{target_lesson['класс']}\n{target_lesson['предмет']}\nКаб. {target_lesson['кабинет']}"
						ws[f'{col}{row}'] = text
						total_hours += 1
					else:
						ws[f'{col}{row}'] = "-"

					ws[f'{col}{row}'].border = border
					ws[f'{col}{row}'].alignment = center_align
					ws[f'{col}{row}'].font = Font(name='Calibri', size=11)
					if day_idx % 2 == 1:
						ws[f'{col}{row}'].fill = alt_fill

				ws.row_dimensions[row].height = 55
				row += 1

			ws[f'A{row}'] = f"Всего часов в неделю: {total_hours}"
			ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
			row += 1
			ws[f'A{row}'] = f"Окна (пропуски): {gaps}"
			ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
			row += 2

		ws.column_dimensions['A'].width = 15
		for col in range(2, 9):
			ws.column_dimensions[get_column_letter(col)].width = 18

	# ===== ЛИСТ 3: ЗАГРУЖЕННОСТЬ КАБИНЕТОВ =====
	def create_rooms_sheet(wb, schedule, days):
		ws = wb.create_sheet("Загруженность кабинетов", 3)
		ws['A1'] = "ЗАГРУЖЕННОСТЬ КАБИНЕТОВ"
		ws['A1'].font = Font(name='Calibri', bold=True, size=14, color="4472C4")
		ws['A1'].alignment = center_align
		ws.merge_cells('A1:H1')
		ws.row_dimensions[1].height = 22

		room_usage = defaultdict(lambda: {day: 0 for day in days})
		rooms_set = set()

		for class_name, class_schedule in schedule.items():
			for day, lessons in class_schedule.items():
				for lesson in lessons:
					if lesson is not None:
						room = lesson.get('кабинет', '101')
						if room and room != 'N/A':
							rooms_set.add(str(room))
							room_usage[str(room)][day] += 1

		sorted_rooms = sorted(rooms_set)
		row = 3

		ws[f'A{row}'] = "Кабинет"
		ws[f'A{row}'].font = header_font
		ws[f'A{row}'].fill = header_fill
		ws[f'A{row}'].alignment = center_align
		ws[f'A{row}'].border = border

		for day_idx, day in enumerate(days):
			col = get_column_letter(day_idx + 2)
			ws[f'{col}{row}'] = day[:2]
			ws[f'{col}{row}'].font = header_font
			ws[f'{col}{row}'].fill = header_fill
			ws[f'{col}{row}'].alignment = center_align
			ws[f'{col}{row}'].border = border

		col = get_column_letter(len(days) + 2)
		ws[f'{col}{row}'] = "Всего"
		ws[f'{col}{row}'].font = header_font
		ws[f'{col}{row}'].fill = header_fill
		ws[f'{col}{row}'].alignment = center_align
		ws[f'{col}{row}'].border = border

		col = get_column_letter(len(days) + 3)
		ws[f'{col}{row}'] = "%"
		ws[f'{col}{row}'].font = header_font
		ws[f'{col}{row}'].fill = header_fill
		ws[f'{col}{row}'].alignment = center_align
		ws[f'{col}{row}'].border = border
		row += 1

		max_lessons_per_day = 7
		total_usage_all = 0
		rooms_used = 0

		for room in sorted_rooms:
			ws[f'A{row}'] = str(room)
			ws[f'A{row}'].border = border
			ws[f'A{row}'].alignment = center_align
			ws[f'A{row}'].font = Font(name='Calibri', size=11)

			daily_usage = []
			for day in days:
				usage = room_usage[room].get(day, 0)
				daily_usage.append(usage)

			total_usage = sum(daily_usage)
			total_usage_all += total_usage

			if total_usage > 0:
				rooms_used += 1

			percentage = (total_usage / (len(days) * max_lessons_per_day)) * 100 if len(days) > 0 else 0

			if percentage <= 25:
				fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
			elif percentage <= 50:
				fill_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
			elif percentage <= 75:
				fill_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
			else:
				fill_color = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

			for day_idx, usage in enumerate(daily_usage):
				col = get_column_letter(day_idx + 2)
				ws[f'{col}{row}'] = usage
				ws[f'{col}{row}'].border = border
				ws[f'{col}{row}'].alignment = center_align
				ws[f'{col}{row}'].fill = fill_color
				ws[f'{col}{row}'].font = Font(name='Calibri', size=11)

			col = get_column_letter(len(days) + 2)
			ws[f'{col}{row}'] = total_usage
			ws[f'{col}{row}'].border = border
			ws[f'{col}{row}'].alignment = center_align
			ws[f'{col}{row}'].fill = fill_color
			ws[f'{col}{row}'].font = Font(name='Calibri', size=11)

			col = get_column_letter(len(days) + 3)
			ws[f'{col}{row}'] = f"{percentage:.0f}%"
			ws[f'{col}{row}'].border = border
			ws[f'{col}{row}'].alignment = center_align
			ws[f'{col}{row}'].fill = fill_color
			ws[f'{col}{row}'].font = Font(name='Calibri', size=11)
			row += 1

		row += 1
		avg_percentage = (total_usage_all / (len(sorted_rooms) * len(days) * max_lessons_per_day)) * 100 if len(sorted_rooms) > 0 and len(days) > 0 else 0

		room_percentages = {}
		for room in sorted_rooms:
			usage = sum(room_usage[room].values())
			percentage = (usage / (len(days) * max_lessons_per_day)) * 100 if len(days) > 0 else 0
			room_percentages[room] = percentage

		max_room = max(room_percentages, key=room_percentages.get) if room_percentages else "N/A"
		min_room = min(room_percentages, key=room_percentages.get) if room_percentages else "N/A"

		ws[f'A{row}'] = f"Всего кабинетов: {len(sorted_rooms)}"
		ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
		row += 1
		ws[f'A{row}'] = f"Используемых: {rooms_used}"
		ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
		row += 1
		ws[f'A{row}'] = f"Пустых: {len(sorted_rooms) - rooms_used}"
		ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
		row += 1
		ws[f'A{row}'] = f"Средняя загруженность: {avg_percentage:.1f}%"
		ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
		row += 1
		ws[f'A{row}'] = f"Самый нагруженный: Каб. {max_room} ({room_percentages.get(max_room, 0):.0f}%)"
		ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)
		row += 1
		ws[f'A{row}'] = f"Самый свободный: Каб. {min_room} ({room_percentages.get(min_room, 0):.0f}%)"
		ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)

		ws.column_dimensions['A'].width = 15
		for col in range(2, len(days) + 4):
			ws.column_dimensions[get_column_letter(col)].width = 12

	create_report_sheet(wb, schedule, fitness, conflicts)
	create_classes_sheet(wb, schedule, days)
	create_teachers_sheet(wb, schedule, days)
	create_rooms_sheet(wb, schedule, days)
	return wb

# ==================== ЭКСПОРТ В PDF ====================

def export_to_pdf(schedule, fitness, conflicts):
    """Экспортировать расписание в PDF"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    buffer = BytesIO()

    # Регистрируем русские шрифты
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuBold', 'DejaVuSans-Bold.ttf'))
        font_name = 'DejaVu'
        font_name_bold = 'DejaVuBold'
    except:
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.3 * inch, bottomMargin=0.3 * inch,
                            leftMargin=0.3 * inch, rightMargin=0.3 * inch)
    story = []

    days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']

    for class_name, class_schedule in sorted(schedule.items()):
        # Заголовок класса (простой текст без стилей)
        header_style = ParagraphStyle(
            'ClassHeader',
            fontName=font_name_bold,
            fontSize=12,
            spaceAfter=10,
            textColor=colors.HexColor('#4472C4')
        )
        story.append(Paragraph(f"Класс {class_name}", header_style))

        data = [['День', 'Урок 1', 'Урок 2', 'Урок 3', 'Урок 4', 'Урок 5', 'Урок 6', 'Урок 7']]

        for day in days:
            row = [day]
            lessons = class_schedule.get(day, [])
            if isinstance(lessons, dict):
                lessons = list(lessons.values())

            for period in range(1, 8):
                target_lesson = None
                for lesson in lessons:
                    if lesson is not None:
                        lesson_num = lesson.get('урок') if isinstance(lesson, dict) else getattr(lesson, 'урок', 0)
                        if lesson_num == period:
                            target_lesson = lesson
                            break

                if target_lesson:
                    subj = target_lesson.get('предмет', '-')
                    if not subj or subj == 'N/A':
                        subj = '-'
                    teacher = target_lesson.get('учитель', '-')
                    if not teacher or teacher == 'N/A':
                        teacher = '-'
                    room = target_lesson.get('кабинет', '-')
                    if not room or room == 'N/A':
                        room = '-'

                    text = f"{subj}\n{teacher}\nКаб.{room}"
                    row.append(text)
                else:
                    row.append('-')

            data.append(row)

        table = Table(data, colWidths=[0.8 * inch] + [0.9 * inch] * 7)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 1), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        ]))

        story.append(table)
        story.append(Spacer(1, 0.2 * inch))

    # Статистика
    stats_style = ParagraphStyle(
        'StatsHeader',
        fontName=font_name_bold,
        fontSize=12,
        spaceAfter=10,
        textColor=colors.HexColor('#4472C4')
    )
    story.append(Paragraph("Статистика", stats_style))

    stats_data = [
        ['Параметр', 'Значение'],
        ['Качество расписания', f"{fitness:.1f}%"],
        ['Конфликты учителей', str(conflicts.get('teacher_conflicts', 0))],
        ['Конфликты кабинетов', str(conflicts.get('room_conflicts', 0))],
        ['Нарушения SanPin', str(conflicts.get('sanpin_violations', 0))],
    ]

    stats_table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    story.append(stats_table)
    doc.build(story)
    buffer.seek(0)
    return buffer
# ==================== ROUTES ====================

@app.route('/')
def serve_index():
	"""Главная страница"""
	return send_file(os.path.join(BASE_DIR, 'index.html'))

@app.route('/api/optimize', methods=['POST'])
def optimize():
	"""Запустить оптимизацию"""
	global ga_instance, last_schedule, last_fitness, last_conflicts
	try:
		data = request.json
		classes = data.get('classes', [])
		subjects = data.get('subjects', [])
		teachers = data.get('teachers', [])
		rooms = data.get('rooms', [])

		if not classes or not subjects or not teachers:
			return jsonify({'success': False, 'error': 'Недостаточно данных'}), 400

		ga_instance = GeneticAlgorithm(
			classes=classes,
			subjects=subjects,
			teachers=teachers,
			rooms=rooms,
			generations=100,
			population_size=30,
			mutation_rate=0.2,
			crossover_rate=0.85
		)

		result = ga_instance.run()
		schedule = result.schedule
		fitness = result.fitness
		conflicts = result.conflicts

		schedule_dict = {}
		for class_name, class_schedule in schedule.items():
			schedule_dict[class_name] = {}
			for day, lessons in class_schedule.items():
				schedule_dict[class_name][day] = []
				for lesson in lessons:
					schedule_dict[class_name][day].append({
						'урок': lesson.get('урок', 0),
						'предмет': lesson.get('предмет', 'N/A'),
						'учитель': lesson.get('учитель', 'N/A'),
						'кабинет': str(lesson.get('кабинет', '101'))
					})

		last_schedule = schedule_dict
		last_fitness = fitness
		last_conflicts = conflicts

		total_lessons = sum(len(lessons) for class_schedule in schedule_dict.values() for lessons in class_schedule.values())

		return jsonify({
			'success': True,
			'schedule': schedule_dict,
			'fitness': fitness,
			'total_lessons': total_lessons,
			'conflicts': conflicts
		})

	except Exception as e:
		import traceback
		print(f"Error: {str(e)}")
		print(traceback.format_exc())
		return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/update', methods=['POST'])
def update_schedule():
	"""Обновить расписание после редактирования"""
	global last_schedule, last_fitness, last_conflicts
	try:
		data = request.json
		new_schedule = data.get('schedule')
		if not new_schedule:
			return jsonify({'success': False, 'error': 'Нет данных'}), 400
		last_schedule = new_schedule
		return jsonify({'success': True})
	except Exception as e:
		return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/export/xlsx', methods=['GET'])
def export_xlsx():
	"""Экспортировать в XLSX"""
	if not last_schedule:
		return abort(400)
	try:
		wb = export_to_excel(last_schedule, last_fitness, last_conflicts)
		buffer = BytesIO()
		wb.save(buffer)
		buffer.seek(0)
		return send_file(
			buffer,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			as_attachment=True,
			download_name=f'расписание_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
		)
	except Exception as e:
		import traceback
		print(traceback.format_exc())
		return jsonify({'error': str(e)}), 400

@app.route('/api/export/pdf', methods=['GET'])
def export_pdf():
	"""Экспортировать в PDF"""
	if not last_schedule:
		return abort(400)
	try:
		buffer = export_to_pdf(last_schedule, last_fitness, last_conflicts)
		return send_file(
			buffer,
			mimetype='application/pdf',
			as_attachment=True,
			download_name=f'расписание_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
		)
	except Exception as e:
		return jsonify({'error': str(e)}), 400

@app.route('/api/export/json', methods=['GET'])
def export_json():
	"""Экспортировать в JSON"""
	if not last_schedule:
		return abort(400)
	try:
		buffer = BytesIO()
		buffer.write(json.dumps(last_schedule, ensure_ascii=False, indent=2).encode('utf-8'))
		buffer.seek(0)
		return send_file(
			buffer,
			mimetype='application/json',
			as_attachment=True,
			download_name=f'расписание_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
		)
	except Exception as e:
		return jsonify({'error': str(e)}), 400

@app.route('/api/health', methods=['GET'])
def health():
	"""Проверка здоровья"""
	return jsonify({'status': 'healthy', 'schedule_ready': last_schedule is not None})

if __name__ == '__main__':
	print("=" * 60)
	print("🚀 Schedule Optimizer API v6.0 запущен!")
	print("=" * 60)
	print("📍 URL: http://localhost:5000")
	print("📊 API: http://localhost:5000/api/optimize")
	print("📥 Экспорт:")
	print(" - XLSX: /api/export/xlsx")
	print(" - PDF: /api/export/pdf")
	print(" - JSON: /api/export/json")
	print("=" * 60)
	app.run(debug=True, port=5000)
